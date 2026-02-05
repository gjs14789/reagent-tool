import streamlit as st
import pandas as pd
import io
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==========================================
# 1. 設定：輸入欄位對照 (讀取 Excel 用)
# ==========================================
# 這些是程式運算邏輯需要的欄位 (來源檔必須包含這些)
INPUT_MAPPING = {
    "id": "產品品號",       # 用於判斷 '物料型態' (原庫存狀態)
    "name": "品名",         # 用於判斷 '系列項目' & '項目分類'
    "order": "製令單號",    # 用於判斷 '季度' & '年份'
    "numerator": "已生產量", # 分子
    "denominator": "預計產量" # 分母
}

# ==========================================
# 2. 設定：最終輸出欄位順序 (寫入 Excel 用)
# ==========================================
# 依照您指定的要求排序
FINAL_COLUMNS_ORDER = [
    "index", "製令單別", "單別名稱", "製令單號", "季度", "急料", "開單日期", "列印", "星期", 
    "性質", "狀態碼", "類型", "物料型態", "系列項目", "項目分類", "產品品號", "品名", 
    "規格", "單位", "BOM版次", "預計產量", "已領套數", "產率", "已生產量", "報廢數量", 
    "備註", "BOM日期", "預計開工", "星期2", "預計完工", "星期3", "實際開工", "星期4", 
    "實際完工", "星期5", "確認日", "確認者", "名稱", "生產廠別", "廠別名稱", "入庫庫別", 
    "庫別名稱", "生產線別", "線別名稱", "加工廠商", "廠商名稱", "稅別碼", "稅別名稱", 
    "生管/採購人員", "人員姓名", "幣別", "課稅別", "營業稅率", "價格條件", "付款條件代號", 
    "付款條件名稱", "預計批號", "送貨地址", "匯率", "加工單位", "計劃批號", "母製令單別", 
    "母製令單號", "訂單單別", "訂單單號", "訂單序號", "客戶代號", "客戶簡稱", "客戶單號", 
    "客戶品號", "確認碼", "簽核狀態", "傳送次數", "EBO拋轉狀態", "版次", "專案代號", 
    "專案名稱", "SMES整合", "SMES拋轉紀錄碼", "ISO單號"
]

# ==========================================
# 3. 核心邏輯函式
# ==========================================

def get_stock_status(val):
    """計算物料型態 (原庫存狀態): 取品號第1碼"""
    s = str(val).strip()
    return s if len(s) > 0 else ""

def classify_product(row):
    """計算系列項目(原產品類別) 與 項目分類(原次分類)"""
    p_name = str(row.get(INPUT_MAPPING["name"], "")).lower().strip()
    stock_status = str(row.get("物料型態", "")).lower()
    
    main_cat = "核酸萃取"
    sub_cat = ""

    # 1. 物料型態判斷
    if stock_status != "a":
        return "非試劑類", ""

    # 2. 關鍵字判斷
    if "extraction" in p_name or "cartridge" in p_name:
        main_cat = "核酸萃取"
    elif any(x in p_name for x in ["pockit", "iq", "dntp", "enzyme", "trehalose", "sedingin", "camap"]):
        main_cat = "配方試劑"
    elif "taco" in p_name:
        main_cat = "核酸萃取"
    elif "ivd" in p_name:
        main_cat = "IVD"
    
    # 3. 次分類判斷
    if main_cat == "核酸萃取":
        if "cartridge" in p_name:
            sub_cat = "POCKIT Central (相關)"
        elif "extraction" in p_name:
            sub_cat = "核酸萃取"
        else:
            sub_cat = "核酸萃取" 
    elif main_cat == "配方試劑":
        if any(x in p_name for x in ["enzyme", "dntp", "iq plus", "pockit"]):
            sub_cat = "IQ Plus、POCKIT"
        elif "pockit central" in p_name or "sedingin" in p_name:
            sub_cat = "POCKIT Central"
        elif any(x in p_name for x in ["camap", "iq200", "iq 2000"]):
            sub_cat = "IQ 2000"
        elif "iq real" in p_name:
            sub_cat = "IQ real"

    return main_cat, sub_cat

def get_quarter(order_val):
    """計算季度"""
    try:
        s = str(order_val).strip()
        if len(s) < 6: return ""
        month = int(s[4:6])
        if 1 <= month <= 3: return "Q1"
        if 4 <= month <= 6: return "Q2"
        if 7 <= month <= 9: return "Q3"
        if 10 <= month <= 12: return "Q4"
        return ""
    except:
        return ""

def process_data(df):
    """執行所有資料運算與欄位重整"""
    
    # 1. 產生 Index
    df.reset_index(drop=True, inplace=True)
    df.index += 1
    df['index'] = df.index

    # 2. 檢查必要欄位
    required = list(INPUT_MAPPING.values())
    missing = [col for col in required if col not in df.columns]
    if missing:
        return None, f"❌ 錯誤：Excel 中找不到這些欄位：{missing}。請確認標題列是否正確。"

    # 3. 計算：物料型態
    df['物料型態'] = df[INPUT_MAPPING["id"]].apply(get_stock_status)

    # 4. 計算：系列項目 & 項目分類
    # 分類邏輯依賴 '物料型態'，所以要先算上面
    classification_result = df.apply(classify_product, axis=1)
    df['系列項目'] = [res for res in classification_result]
    df['項目分類'] = [res[1] for res in classification_result]

    # 5. 計算：季度 & 年份 (用於統計)
    df['季度'] = df[INPUT_MAPPING["order"]].apply(get_quarter)
    df['年份'] = df[INPUT_MAPPING["order"]].astype(str).str[:4] # 暫存用於統計，不輸出

    # 6. 計算：產率
    def calc_yield(row):
        try:
            num = float(row.get(INPUT_MAPPING["numerator"], 0))
            den = float(row.get(INPUT_MAPPING["denominator"], 0))
            return num / den if den != 0 else 0
        except:
            return 0
    df['產率'] = df.apply(calc_yield, axis=1)

    # 7. 統計年份 (MsgBox功能)
    stats = df['年份'].value_counts().sort_index().to_dict()

    # 8. 欄位排序與補缺
    # 建立一個只包含目標欄位的 DataFrame，若原資料沒有該欄位則填入空值
    final_df = pd.DataFrame()
    for col in FINAL_COLUMNS_ORDER:
        if col in df.columns:
            final_df[col] = df[col]
        else:
            final_df[col] = "" # 若原檔沒有此欄位，填空白

    return final_df, stats

# ==========================================
# 4. Streamlit 介面與檔案處理
# ==========================================

st.set_page_config(page_title="製造命令處理工具", page_icon="🏭")
st.title("🏭 製造命令單頭資料前處理")
st.markdown("### 說明")
st.info("本工具會保留原始 Excel 所有工作表，並新增一個包含計算結果的工作表。")

uploaded_file = st.file_uploader("請上傳 Excel 檔案", type=["xlsx", "xlsm"])

if uploaded_file:
    try:
        # 使用 openpyxl 載入整個活頁簿 (為了保留原始檔案內容)
        wb = openpyxl.load_workbook(uploaded_file)
        sheet_names = wb.sheetnames
        
        selected_sheet = st.selectbox("請選擇要處理的原始資料工作表：", sheet_names)
        
        if st.button("開始處理"):
            with st.spinner('正在分析與生成報表...'):
                # 為了計算方便，這裡用 pandas 再讀一次資料 (只讀選定的 sheet)
                # header=2 代表 Excel 第 3 列是標題
                df_raw = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=2)
                
                # 執行運算
                result_df, stats = process_data(df_raw.copy())
                
                if result_df is not None:
                    # --- 處理工作表命名 (需求 2) ---
                    base_name = f"{selected_sheet}的處理結果"
                    count = 1
                    new_sheet_name = f"{base_name}({count})"
                    
                    # 檢查名稱是否重複，若重複則數字+1
                    while new_sheet_name in wb.sheetnames:
                        count += 1
                        new_sheet_name = f"{base_name}({count})"
                    
                    # --- 新增工作表並寫入資料 (需求 3) ---
                    ws_new = wb.create_sheet(new_sheet_name)
                    
                    # 將 DataFrame 寫入新的 sheet (含標題)
                    for r in dataframe_to_rows(result_df, index=False, header=True):
                        ws_new.append(r)
                    
                    # --- 設定表格格式 (Table Style) ---
                    # 定義表格範圍 (例如 A1:AC100)
                    max_col_letter = openpyxl.utils.get_column_letter(len(result_df.columns))
                    max_row = len(result_df) + 1 # +1 是標題列
                    table_ref = f"A1:{max_col_letter}{max_row}"
                    
                    # 建立表格物件 (類似 VBA 的 ListObject)
                    tab = Table(displayName=f"Table_{new_sheet_name.replace('(', '_').replace(')', '_')}", ref=table_ref)
                    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    tab.tableStyleInfo = style
                    ws_new.add_table(tab)
                    
                    # --- 設定產率欄位為百分比格式 ---
                    if "產率" in result_df.columns:
                        # 找出產率是第幾欄 (1-based index)
                        yield_col_idx = result_df.columns.get_loc("產率") + 1 
                        yield_col_letter = openpyxl.utils.get_column_letter(yield_col_idx)
                        # 設定整欄格式
                        for cell in ws_new[yield_col_letter]:
                            # 跳過第一列標題
                            if cell.row > 1: 
                                cell.number_format = '0.00%'

                    # --- 存檔準備下載 ---
                    virtual_workbook = io.BytesIO()
                    wb.save(virtual_workbook)
                    virtual_workbook.seek(0)
                    
                    # 顯示成功資訊
                    st.success(f"✅ 處理完成！已新增工作表：`{new_sheet_name}`")
                    st.write("📊 **年度統計：**", stats)
                    
                    st.download_button(
                        label="📥 下載完整 Excel 檔案",
                        data=virtual_workbook,
                        file_name=f"Processed_{uploaded_file.name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.stop()

    except Exception as e:
        st.error(f"發生錯誤：{str(e)}")
        st.error("請確認上傳的是有效的 Excel 檔案。")
